
import io
import base64
import traceback
from typing import List, Optional

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.cm as mcm
import matplotlib.colors as mcolors
from matplotlib.figure import Figure
from matplotlib.backends.backend_agg import FigureCanvasAgg
import numpy as np

from fastapi import FastAPI
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel

import sys, os
sys.path.insert(0, os.path.abspath(os.path.dirname(__file__)))
from pdesolver import PDE, PDES

app = FastAPI(title='PDESsolver API')

_cache: dict = {}

app.add_middleware(
    CORSMiddleware,
    allow_origins=['http://localhost:5173', 'http://localhost:3000'],
    allow_methods=['*'],
    allow_headers=['*'],
)



class PDEConfig(BaseModel):
    equation: str          
    func:     str          
    ic:       str          
    west_bd:  str = 'Dirichlet'
    west_val: str = '0'
    east_bd:  str = 'Neumann'
    east_val: str = '0'
    south_bd: str = 'Neumann'
    south_val: str = '0'
    north_bd: str = 'Neumann'
    north_val: str = '0'


class SimRequest(BaseModel):
    pdes:          List[PDEConfig]
    dimension:     str   = '1D'      
    nx:            int   = 21
    ny:            int   = 21
    tf:            float = 1.0
    nt:            int   = 200
    tol:           float = 1e-6
    method:        str   = 'bdf2'    
    disc_method:   str   = 'central' 
    func_idx:      int   = 0
    plot_mode:     str   = 'final'   


def fig_to_b64(fig: Figure) -> str:
    buf = io.BytesIO()
    fig.savefig(buf, format='png', dpi=130, bbox_inches='tight', facecolor='white')
    buf.seek(0)
    return base64.b64encode(buf.read()).decode()


PLOT_STYLE = {
    'figure.facecolor':  'white',
    'axes.facecolor':    '#f8f9fb',
    'axes.edgecolor':    '#cccccc',
    'axes.labelcolor':   '#222222',
    'axes.labelsize':    13,
    'axes.titlesize':    15,
    'axes.titlecolor':   '#111111',
    'axes.grid':         True,
    'grid.color':        '#e0e0e0',
    'grid.linewidth':    0.7,
    'xtick.color':       '#333333',
    'xtick.labelsize':   11,
    'ytick.color':       '#333333',
    'ytick.labelsize':   11,
    'text.color':        '#222222',
    'lines.linewidth':   2.4,
    'legend.fontsize':   11,
    'legend.facecolor':  'white',
    'legend.edgecolor':  '#cccccc',
    'legend.labelcolor': '#222222',
}


def make_fig(w=7, h=4):
    with plt.rc_context(PLOT_STYLE):
        fig = Figure(figsize=(w, h), facecolor='white')
    FigureCanvasAgg(fig)
    return fig


def plot_final(hist, funcs, disc_n, pdes_list, func_idx):
    a, b = pdes_list[0].ivar_boundary[0]
    x    = np.linspace(a, b, disc_n[0])
    n    = len(funcs)
    colors = ['#4fc3f7', '#81c784', '#ffb74d', '#f48fb1',
              '#ce93d8', '#80cbc4', '#fff176', '#ef9a9a']

    fig = make_fig()
    ax  = fig.add_subplot(111, facecolor='#1a1d27')
    for i, f in enumerate(funcs):
        ax.plot(x, np.array(hist[i][-1]),
                color=colors[i % len(colors)],
                linewidth=2.2, label=f)
    ax.set_xlabel('x'); ax.set_ylabel('u(x, tf)')
    ax.set_title('Perfil final — t = tf', color='#e2e8f4', fontsize=13, pad=10)
    ax.legend(facecolor='#1a1d27', edgecolor='#2e3347', labelcolor='#c9d1e0')
    fig.tight_layout()
    return fig_to_b64(fig)


def plot_xt(hist, funcs, disc_n, pdes_list, func_idx, tf):
    a, b   = pdes_list[0].ivar_boundary[0]
    matriz = np.array(hist[func_idx])
    fig = make_fig()
    ax  = fig.add_subplot(111, facecolor='#1a1d27')
    im  = ax.imshow(matriz.T, aspect='auto', origin='lower',
                    extent=[0, tf, a, b], cmap='plasma')
    fig.colorbar(im, ax=ax)
    ax.set_xlabel('t'); ax.set_ylabel('x')
    ax.set_title(f'{funcs[func_idx]} — espaço × tempo',
                 color='#e2e8f4', fontsize=13, pad=10)
    fig.tight_layout()
    return fig_to_b64(fig)


def plot_profiles(hist, funcs, disc_n, pdes_list, func_idx, tf, n=8):
    a, b  = pdes_list[0].ivar_boundary[0]
    x     = np.linspace(a, b, disc_n[0])
    n_p   = len(hist[func_idx])
    idxs  = np.linspace(0, n_p-1, n, dtype=int)
    cmap_ = mcm.get_cmap('cool')

    fig = make_fig()
    ax  = fig.add_subplot(111, facecolor='#1a1d27')
    for k, idx in enumerate(idxs):
        t_val = tf * idx / max(n_p-1, 1)
        ax.plot(x, np.array(hist[func_idx][idx]),
                color=cmap_(k / max(n-1, 1)),
                linewidth=1.8, label=f't={t_val:.2f}', alpha=0.9)
    ax.set_xlabel('x'); ax.set_ylabel(funcs[func_idx])
    ax.set_title(f'{funcs[func_idx]} — evolução temporal',
                 color='#e2e8f4', fontsize=13, pad=10)
    ax.legend(fontsize=7, facecolor='#1a1d27',
              edgecolor='#2e3347', labelcolor='#c9d1e0')
    sm = mcm.ScalarMappable(cmap='cool',
                            norm=mcolors.Normalize(vmin=0, vmax=tf))
    sm.set_array([])
    fig.colorbar(sm, ax=ax, label='t')
    fig.tight_layout()
    return fig_to_b64(fig)


def plot_2d_heatmap(hist, funcs, disc_n, func_idx):
    nx, ny = disc_n
    data   = np.array(hist[func_idx][-1]).reshape(nx, ny)
    X, Y   = np.meshgrid(np.linspace(0,1,nx), np.linspace(0,1,ny), indexing='ij')
    fig = make_fig(6, 5)
    ax  = fig.add_subplot(111, facecolor='#1a1d27')
    c   = ax.contourf(X, Y, data, levels=30, cmap='RdYlBu_r')
    fig.colorbar(c, ax=ax)
    ax.set_xlabel('x'); ax.set_ylabel('y')
    ax.set_title(f'{funcs[func_idx]} — estado final',
                 color='#e2e8f4', fontsize=13, pad=10)
    fig.tight_layout()
    return fig_to_b64(fig)


def plot_2d_surface(hist, funcs, disc_n, func_idx):
    nx, ny = disc_n
    data   = np.array(hist[func_idx][-1]).reshape(nx, ny)
    X, Y   = np.meshgrid(np.linspace(0,1,nx), np.linspace(0,1,ny), indexing='ij')
    fig = make_fig(7, 5)
    ax  = fig.add_subplot(111, projection='3d', facecolor='#1a1d27')
    ax.plot_surface(X, Y, data, cmap='RdYlBu_r', alpha=0.92)
    ax.set_xlabel('x'); ax.set_ylabel('y'); ax.set_zlabel(funcs[func_idx])
    ax.set_title(f'{funcs[func_idx]} — superfície',
                 color='#e2e8f4', fontsize=12, pad=10)
    fig.tight_layout()
    return fig_to_b64(fig)


@app.get('/')
def root():
    return {'status': 'PDESsolver API running'}


@app.post('/solve_json')
def solve_json(payload: dict):
    try:
        from backend.solvers import dispatch
        import time
        t0 = time.perf_counter()
        result = dispatch(payload)
        result["meta"]["elapsed_ms"] = int((time.perf_counter() - t0) * 1000)
        return result
    except Exception as e:
        traceback.print_exc()
        return {'ok': False, 'error': str(e)}


@app.post('/solve')
def solve(req: SimRequest):
    try:
        is2d   = req.dimension == '2D'
        disc_n = [req.nx, req.ny] if is2d else [req.nx]

        pde_list = []
        for p in req.pdes:
            sp_var = ['x','y'] if is2d else ['x']
            bounds = [(0,1),(0,1)] if is2d else [(0,1)]
            letra  = p.func
            pde_list.append(PDE.PDE(
                f'd{letra}/dt = {p.equation}',
                letra, sp_var, ['t'], bounds,
                expr_ic       = p.ic or '0',
                west_bd       = p.west_bd,
                west_func_bd  = p.west_val or '0',
                east_bd       = p.east_bd,
                east_func_bd  = p.east_val or '0',
                north_bd      = p.north_bd if is2d else 'Neumann',
                north_func_bd = p.north_val if is2d else '0',
                south_bd      = p.south_bd if is2d else 'Neumann',
                south_func_bd = p.south_val if is2d else '0',
            ))

        sim = PDES(pde_list, disc_n)
        sim.discretize(method=req.disc_method)
        sim.solve(method=req.method, tf=req.tf, nt=req.nt, tol=req.tol)

        _, hist   = sim.results
        func_idx  = req.func_idx
        mode      = req.plot_mode

        if is2d:
            b64 = (plot_2d_surface(hist, sim.funcs, disc_n, func_idx)
                   if mode == 'surface2d'
                   else plot_2d_heatmap(hist, sim.funcs, disc_n, func_idx))
        else:
            if mode == 'xt':
                b64 = plot_xt(hist, sim.funcs, disc_n, pde_list, func_idx, req.tf)
            elif mode == 'profiles':
                b64 = plot_profiles(hist, sim.funcs, disc_n,
                                    pde_list, func_idx, req.tf)
            else:
                b64 = plot_final(hist, sim.funcs, disc_n, pde_list, func_idx)

        _cache['hist']      = hist
        _cache['funcs']     = sim.funcs
        _cache['disc_n']    = disc_n
        _cache['pde_list']  = pde_list
        _cache['is_2d']     = is2d
        _cache['tf']        = req.tf

        return {
            'ok':    True,
            'image': b64,
            'funcs': sim.funcs,
        }

    except Exception as e:
        traceback.print_exc()
        return {'ok': False, 'error': str(e)}


class ReplotRequest(BaseModel):
    func_idx:  int = 0
    plot_mode: str = 'final'


@app.post('/replot')
def replot(req: ReplotRequest):
    if not _cache:
        return {'ok': False, 'error': 'Nenhuma simulação em cache. Rode /solve primeiro.'}
    try:
        hist     = _cache['hist']
        funcs    = _cache['funcs']
        disc_n   = _cache['disc_n']
        pde_list = _cache['pde_list']
        is_2d    = _cache['is_2d']
        tf       = _cache['tf']
        fidx     = req.func_idx
        mode     = req.plot_mode

        if is_2d:
            b64 = (plot_2d_surface(hist, funcs, disc_n, fidx)
                   if mode == 'surface2d'
                   else plot_2d_heatmap(hist, funcs, disc_n, fidx))
        else:
            if mode == 'xt':
                b64 = plot_xt(hist, funcs, disc_n, pde_list, fidx, tf)
            elif mode == 'profiles':
                b64 = plot_profiles(hist, funcs, disc_n, pde_list, fidx, tf)
            else:
                b64 = plot_final(hist, funcs, disc_n, pde_list, fidx)

        return {'ok': True, 'image': b64, 'funcs': funcs}
    except Exception as e:
        traceback.print_exc()
        return {'ok': False, 'error': str(e)}




@app.post('/download_xlsx')
def download_xlsx(req: SimRequest):
    import io
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return {'ok': False, 'error': 'openpyxl não instalado. pip install openpyxl'}

    result = solve(req)
    if not result.get('ok'):
        return result

    is2d   = req.dimension == '2D'
    disc_n = [req.nx, req.ny] if is2d else [req.nx]
    pde_list = []
    for p in req.pdes:
        sp_var = ['x','y'] if is2d else ['x']
        bounds = [(0,1),(0,1)] if is2d else [(0,1)]
        letra  = p.func
        pde_list.append(PDE.PDE(
            f'd{letra}/dt = {p.equation}',
            letra, sp_var, ['t'], bounds,
            expr_ic=p.ic or '0',
            west_bd=p.west_bd, west_func_bd=p.west_val or '0',
            east_bd=p.east_bd, east_func_bd=p.east_val or '0',
            north_bd=p.north_bd if is2d else 'Neumann',
            north_func_bd=p.north_val if is2d else '0',
            south_bd=p.south_bd if is2d else 'Neumann',
            south_func_bd=p.south_val if is2d else '0',
        ))

    sim = PDES(pde_list, disc_n)
    sim.discretize(method=req.disc_method)
    sim.solve(method=req.method, tf=req.tf, nt=req.nt, tol=req.tol)
    _, hist = sim.results

    wb = openpyxl.Workbook()

    ws_cfg = wb.active
    ws_cfg.title = 'Configuração'
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='3949AB')
    headers = ['Parâmetro', 'Valor']
    for c, h in enumerate(headers, 1):
        cell = ws_cfg.cell(1, c, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    cfg_rows = [
        ('Dimensão', req.dimension),
        ('nx', req.nx), ('ny', req.ny if is2d else '-'),
        ('tf', req.tf), ('nt', req.nt), ('tol', req.tol),
        ('Solver', req.method), ('Discretização', req.disc_method),
    ]
    for i, (k, v) in enumerate(cfg_rows, 2):
        ws_cfg.cell(i, 1, k)
        ws_cfg.cell(i, 2, str(v))
    ws_cfg.column_dimensions['A'].width = 18
    ws_cfg.column_dimensions['B'].width = 20

    nx = req.nx
    ny = req.ny if is2d else 1
    dt = req.tf / req.nt
    n_passos = len(hist[0])
    t_values = [round(i * dt, 6) for i in range(n_passos)]
    x_values = [round(i / (nx - 1), 6) for i in range(nx)]

    for fi, func_name in enumerate(sim.funcs):
        ws = wb.create_sheet(title=f'{func_name} — final')

        if not is2d:
            ws.cell(1, 1, 'x / t')
            for ti, t in enumerate(t_values):
                cell = ws.cell(1, ti + 2, t)
                cell.font = Font(bold=True)
                cell.fill = PatternFill('solid', fgColor='E8EAF6')
            for xi, x in enumerate(x_values):
                ws.cell(xi + 2, 1, x)
                for ti in range(n_passos):
                    ws.cell(xi + 2, ti + 2, round(float(hist[fi][ti][xi]), 8))
        else:
            data = np.array(hist[fi][-1]).reshape(nx, ny)
            ws.cell(1, 1, f'{func_name} (t=tf)')
            for j in range(ny):
                ws.cell(1, j + 2, round(j / max(ny-1,1), 4))
            for i in range(nx):
                ws.cell(i + 2, 1, round(i / max(nx-1,1), 4))
                for j in range(ny):
                    ws.cell(i + 2, j + 2, round(float(data[i, j]), 8))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    b64 = base64.b64encode(buf.read()).decode()
    return {'ok': True, 'xlsx': b64, 'filename': 'pdesolver_dados.xlsx'}

if __name__ == '__main__':
    import uvicorn
    uvicorn.run('backend:app', host='localhost', port=8000, reload=True)